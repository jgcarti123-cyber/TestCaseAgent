#!/usr/bin/env python3
"""Project timeline sync - past commits + forward-looking roadmap, as .xlsx.

Produces Docs/project_timeline.xlsx so a future Claude session (or a
human) can reconstruct "what's been built, when, and what's left" in
one read, instead of re-deriving it from `git log` plus diffing every
Docs/*.md file. Three sheets, each regenerated from a live source -
nothing here is hand-maintained, so it can't go stale the way a manual
changelog would:

  1. Commit History      <- `git log` (this repo's own history)
  2. Roadmap - Requirements Items <- Docs/requirements.md's numbered
                              item headers ("## N. Title - status (date)")
  3. Roadmap - Tools Needed        <- Docs/tools.md's "Needed, not yet
                              built" markdown table

Re-run this any time the picture has moved:

    python Scripts/project_timeline_sync.py

Deliberately not gitignored (unlike Signal_Catalogs/, Existing_TestCases/,
etc.) - this file contains no proprietary TML data, only commit metadata
and text already present in the tracked Docs/*.md files.
"""

import re
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
REQUIREMENTS_PATH = REPO_ROOT / "Docs" / "requirements.md"
TOOLS_PATH = REPO_ROOT / "Docs" / "tools.md"
OUTPUT_PATH = REPO_ROOT / "Docs" / "project_timeline.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = "4472C4"


def _style_header_row(ws, ncols):
    from openpyxl.styles import PatternFill

    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# ---- Sheet 1: Commit History ---------------------------------------------


def _parse_git_log():
    """Every commit on the current branch: hash, date, author, message, stat."""

    raw = subprocess.run(
        [
            "git",
            "log",
            "--pretty=format:COMMIT|%H|%h|%ad|%an|%s",
            "--date=format:%Y-%m-%d %H:%M",
            "--shortstat",
        ],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout

    commits = []
    current = None
    stat_re = re.compile(
        r"(\d+) files? changed(?:, (\d+) insertions?\(\+\))?(?:, (\d+) deletions?\(-\))?"
    )

    for line in raw.splitlines():
        if line.startswith("COMMIT|"):
            if current is not None:
                commits.append(current)
            _, full_hash, short_hash, date, author, message = line.split("|", 5)
            current = {
                "hash": short_hash,
                "full_hash": full_hash,
                "date": date,
                "author": author,
                "message": message,
                "files_changed": 0,
                "insertions": 0,
                "deletions": 0,
            }
        elif line.strip() and current is not None:
            m = stat_re.search(line)
            if m:
                current["files_changed"] = int(m.group(1))
                current["insertions"] = int(m.group(2) or 0)
                current["deletions"] = int(m.group(3) or 0)

    if current is not None:
        commits.append(current)

    return commits


def _write_commit_history_sheet(wb):
    ws = wb.active
    ws.title = "Commit History"

    columns = ["Date", "Commit", "Author", "Message", "Files Changed", "Insertions", "Deletions"]
    ws.append(columns)

    for commit in _parse_git_log():
        ws.append(
            [
                commit["date"],
                commit["hash"],
                commit["author"],
                commit["message"],
                commit["files_changed"],
                commit["insertions"],
                commit["deletions"],
            ]
        )

    _style_header_row(ws, len(columns))
    _autosize(ws, [17, 10, 20, 75, 14, 12, 12])
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")


# ---- Sheet 2: Roadmap - Requirements Items --------------------------------

_REQ_HEADER_RE = re.compile(r"^##\s+(\d+)\.\s+(.+?)\s+—\s+(.+)$")


def _parse_requirements_md():
    """Docs/requirements.md's numbered item headers, e.g.
    '## 4. Requirement traceability data — 🟡 LIVE DATA PULLED FOR NIO-F001, GRANULARITY GAP FOUND (2026-08-06)'
    Body text is everything until the next '## N.' header, kept as a
    trimmed first-paragraph summary rather than the full section (full
    detail stays in requirements.md itself - this sheet is a map to it,
    not a replacement for it).
    """

    text = REQUIREMENTS_PATH.read_text()
    lines = text.splitlines()

    items = []
    current = None
    body_lines = []

    def flush():
        if current is not None:
            body = " ".join(l.strip() for l in body_lines if l.strip() and not l.strip().startswith("##"))
            current["summary"] = body[:500]
        if current is not None:
            items.append(current)

    for line in lines:
        m = _REQ_HEADER_RE.match(line.strip())
        if m:
            flush()
            body_lines = []
            current = {"item": int(m.group(1)), "title": m.group(2), "status": m.group(3)}
        else:
            body_lines.append(line)
    flush()

    return items


def _write_requirements_roadmap_sheet(wb):
    ws = wb.create_sheet("Roadmap - Requirements Items")
    columns = ["Item #", "Title", "Status (as written in requirements.md)", "Summary (first paragraph, see requirements.md for full detail)"]
    ws.append(columns)

    for item in _parse_requirements_md():
        ws.append([item["item"], item["title"], item["status"], item.get("summary", "")])

    _style_header_row(ws, len(columns))
    _autosize(ws, [8, 32, 45, 90])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---- Sheet 3: Roadmap - Tools Needed --------------------------------------


def _parse_tools_needed_table():
    """The 'Needed, not yet built' markdown table in Docs/tools.md."""

    text = TOOLS_PATH.read_text()
    section_match = re.search(r"## Needed, not yet built\n\n(.+?)(?:\n\n##|\Z)", text, re.DOTALL)
    if not section_match:
        return []

    rows = []
    for line in section_match.group(1).splitlines():
        line = line.strip()
        if not line.startswith("|") or set(line.replace("|", "").strip()) <= {"-", " "}:
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if cells == ["Tool", "Purpose", "Priority"]:
            continue
        if len(cells) == 3:
            rows.append(cells)

    return rows


def _write_tools_roadmap_sheet(wb):
    ws = wb.create_sheet("Roadmap - Tools Needed")
    columns = ["Tool", "Purpose", "Priority"]
    ws.append(columns)

    for tool, purpose, priority in _parse_tools_needed_table():
        # Strip markdown bold/code markers so the cell reads as plain text.
        clean = lambda s: re.sub(r"[*`]", "", s)
        ws.append([clean(tool), clean(purpose), clean(priority)])

    _style_header_row(ws, len(columns))
    _autosize(ws, [30, 90, 40])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    wb = openpyxl.Workbook()
    _write_commit_history_sheet(wb)
    _write_requirements_roadmap_sheet(wb)
    _write_tools_roadmap_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}", file=sys.stderr)


if __name__ == "__main__":
    main()
