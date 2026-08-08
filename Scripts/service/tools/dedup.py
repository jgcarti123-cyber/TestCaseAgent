"""Exact-field-match dedup - Docs/definition_of_done.md Tier 1 gate #6.

Match key: parent_id + requirement_id + test_set_category + primary
trigger signal (first row of can_signals_referenced). Category-scoped
per Docs/guardrails.md #6: only Edge Case - */User-Journey matches
block release; Happy Path/Negative Case matches are expected regression
coverage and are logged, never blocked (see config.DEDUP_BLOCKING_CATEGORIES).

Scope honestly, not silently: this only checks the current batch and
Generated_TestCases/. It deliberately does NOT check Existing_TestCases/
- per CLAUDE.md, all 2,704 rows there have an empty Requirement ID and
an empty Test Set Category, so an exact match on those two fields could
never fire against that file (or worse, would need to treat "both
empty" as a match, which would false-positive against nearly every
historical row). It also does not check Jira
(Traceability/requirement_traceability.json) - Jira's traceability is
feature-level, not per-requirement-line (see Docs/traceability.md), so
this exact-match method doesn't apply there either. Both gaps are the
documented "Dedup/similarity tool" future upgrade in Docs/tools.md
(embedding similarity against real summaries) - not something this
exact-match method can honestly claim to cover. Callers get told this
via `unchecked_sources`, not a silent pass.
"""

import glob
from functools import lru_cache

import openpyxl

from ..config import DEDUP_BLOCKING_CATEGORIES, GENERATED_TEST_CASES_DIR

_MATCH_COLUMNS = ("Parent ID", "Requirement ID", "Test Set Category")


def _match_key(parent_id: str, requirement_id: str, test_set_category: str, primary_trigger_signal: str) -> tuple:
    return (parent_id, requirement_id, test_set_category, primary_trigger_signal)


def _primary_trigger_signal(can_signals_referenced: str) -> str:
    """First row of the pipe-delimited can_signals_referenced column."""
    first_line = (can_signals_referenced or "").strip().splitlines()[0] if can_signals_referenced else ""
    return first_line.split("|")[0].strip()


@lru_cache(maxsize=1)
def _load_generated_test_cases() -> list[tuple[str, int, tuple]]:
    """Every row already written to Generated_TestCases/, as (file, row_num, match_key)."""

    rows = []
    for path in sorted(glob.glob(str(GENERATED_TEST_CASES_DIR / "*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            col = {name: header.index(name) for name in (*_MATCH_COLUMNS, "CAN Signals Referenced")}
        except ValueError:
            # Sheet doesn't have the expected new-schema columns yet - skip
            # rather than crash the whole dedup check on one malformed file.
            continue

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            parent_id = row[col["Parent ID"]] or ""
            requirement_id = row[col["Requirement ID"]] or ""
            test_set_category = row[col["Test Set Category"]] or ""
            primary_trigger_signal = _primary_trigger_signal(row[col["CAN Signals Referenced"]])
            if not (parent_id and requirement_id and test_set_category and primary_trigger_signal):
                continue
            key = _match_key(parent_id, requirement_id, test_set_category, primary_trigger_signal)
            rows.append((path, row_num, key))
    return rows


def check_dedup(candidate: dict, batch: list[dict]) -> dict:
    """candidate/batch items: {parent_id, requirement_id, test_set_category, primary_trigger_signal}."""

    candidate_key = _match_key(
        candidate["parent_id"],
        candidate["requirement_id"],
        candidate["test_set_category"],
        candidate["primary_trigger_signal"],
    )

    matches = []

    for i, other in enumerate(batch):
        other_key = _match_key(
            other["parent_id"], other["requirement_id"], other["test_set_category"], other["primary_trigger_signal"]
        )
        if other_key == candidate_key:
            matches.append({"source": "batch", "location": f"batch[{i}]"})

    for path, row_num, key in _load_generated_test_cases():
        if key == candidate_key:
            matches.append({"source": "generated_test_cases", "location": f"{path}:row {row_num}"})

    is_duplicate = len(matches) > 0
    blocking = is_duplicate and candidate["test_set_category"] in DEDUP_BLOCKING_CATEGORIES

    return {
        "is_duplicate": is_duplicate,
        "blocking": blocking,
        "matches": matches,
        "unchecked_sources": [
            "Existing_TestCases/ - Requirement ID and Test Set Category are empty on all 2,704 historical "
            "rows, so exact-field match cannot meaningfully apply (see Scripts/service/tools/dedup.py docstring)",
            "Traceability/requirement_traceability.json (Jira) - traceability there is feature-level, not "
            "per-requirement-line; exact-field match doesn't apply. Needs the embedding-similarity tool "
            "described in Docs/tools.md",
        ],
    }
